import pandas as pd
import numpy as np
from openpyxl import load_workbook
from pathlib import Path

from question import Question
import config


class Questions:
    """Class represents a set of questions, as read from survey deinition document.
    Contains information about the question set (e.g. all positively scored) questions as well as a list of
    Question objects.

    init takes a dataframe of the current periods questions. If there are and historical quesstions defined in the
    definition document, these are optionally passed in as dfs.
    """

    def __init__(self, period_questions, **previous_period_questions):

        self._df = period_questions

        self._historic_dfs = previous_period_questions if previous_period_questions else None
        self._historic_indicies = self._get_historic_df_indicies() if previous_period_questions else None

        self._question_list = self._generate_question_list()

        # Properties to generate only if accessed
        self._scored_questions = None
        self._single_response_questions = None
        self._multi_response_questions = None
        self._text_response_questions = None
        self._all_question_columns = None
        self._questions_with_historic = None

    @classmethod
    def from_file(cls, questions_path):
        """ Loads the question definition document from the path, can handle excels and csvs.
         If Excel, if theres only one sheet load that. If multiple sheetsload parse the one with the
         correct sheet name defined in config for main questions.

         If there are sheets defining historic questions (ends with a period identifier),
         parse these and return ther objects"""

        questions_path = Path(questions_path)
        kwargs_dict = {}

        if questions_path.suffix.lower() in config.EXCEL_EXTS:
            wb = load_workbook(questions_path)

            question_sheet_name = config.P_QUESTION_SHEET if len(wb.sheetnames) > 1 else wb.sheetnames[0]
            period_questions = pd.read_excel(questions_path, sheet_name=question_sheet_name)

            if len(wb.sheetnames) > 1:
                for sheet in wb.sheetnames:
                    if sheet[:2] == config.PERIOD_IDENTIFIER:
                        kwargs_dict[sheet] = pd.read_excel(questions_path, sheet_name=sheet)

        elif questions_path.suffix.lower() == '.csv':
            period_questions = pd.read_csv(questions_path, encoding='utf-8')

        return cls(period_questions, **kwargs_dict)

    @property
    def question_list(self):
        return self._question_list

    # Propteries only generated if requested, not really needed tbh should just always populate.
    @property
    def scored_questions(self):
        """returns a list of all scored questions"""
        if not self._scored_questions:
            self._scored_questions = [q for q in self._question_list if q.scored]
        return self._scored_questions

    @property
    def questions_with_historic(self):
        """returns a list of all questions which have matched historic versions"""
        if not self._questions_with_historic:
            self._questions_with_historic = [q for q in self._question_list if len(q.historic_questions) > 0]
        return self._questions_with_historic

    @property
    def single_response_questions(self):
        """returns a list of all questions which are single response"""
        if not self._single_response_questions:
            self._single_response_questions = self._create_type_lists(config.SINGLE_CODE)
        return self._single_response_questions

    @property
    def multi_response_questions(self):
        """returns a list of all questions which are multi response"""
        if not self._multi_response_questions:
            self._multi_response_questions = self._create_type_lists(config.MULTI_CODE)
        return self._multi_response_questions

    @property
    def text_response_questions(self):
        """returns a list of all questions which are test responses"""
        if not self._text_response_questions:
            self._text_response_questions = self._create_type_lists(config.TEXT_CODE)
        return self._text_response_questions

    @property
    def all_question_columns(self):
        """returns a list of all columns covered by this set of questions (i.e. all multi-columns etc)"""
        if not self._all_question_columns:
            all_question_columns = [q.question_columns for q in self._question_list]
            self._all_question_columns = [item for sublist in all_question_columns for item in sublist]
        return self._all_question_columns

    def _create_maps(self, q_row, response_indices, score_indices):
        """Returns both the list of response options and the score map for a question.
        Accepts a single row from the survey definition sheet, as well as the indicies of where
        the response options and score information is"""
        return {'response_options': self._create_map(q_row, response_indices),
                'score_map': self._create_map(q_row, score_indices)}

    @staticmethod
    def _create_map(q_row, indices):
        """return a dictiopnary of the response code and the response text for a question.
        Accepts a single row from the survey definition sheet,
        as well as the indicies of where the response options are

        e.g.
        {1: 'Yes', 2: 'No'}

        """
        mydict = {}
        for i, x in enumerate(q_row[(indices[0]):(indices[-1]) + 1]):
            if isinstance(x, str) or not np.isnan(x):
                x = int(x) if isinstance(x, float) else x
                mydict[i] = x
        return mydict

    @staticmethod
    def _create_indicies(df, index_prefix):
        """Returns list of indicies (i.e. locations) of set of columns with the given prefix e.g. R_"""
        return [i for i, elem in enumerate(list(df.columns)) if index_prefix in elem[0:2]]

    def _get_score_response_indicies(self, df):
        """Returns dict of both score and response column indicies

        e.g:

        {'score_indices': [1,2,3], 'response_indices': [4,5,6]}

        """
        return {'score_indices': self._create_indicies(df, config.HEAD_SCORE_PREFIX),
                'response_indices': self._create_indicies(df, config.HEAD_RESPONSE_PREFIX)}

    def _get_historic_df_indicies(self):
        """Returns dict of both score and response column indicies for historic questions

        e.g.

        {'P-1': {'score_indices': [1,2,3], 'response_indices': [4,5,6]}}

        """
        return {period: self._get_score_response_indicies(historic_df) for period, historic_df in
                self._historic_dfs.items()}

    def _generate_question_list(self):
        indices = self._get_score_response_indicies(self._df)
        questions_list = []
        for index, q_row in self._df.iterrows():
            question_object = self._create_question(q_row, indices)

            if question_object.scored and self._historic_dfs:
                self._create_historic_questions(question_object)

            questions_list.append(question_object)

        return questions_list

    def _create_question(self, q_row, indices):
        """Creates a question object from a row of the survey definiution doc.
        indicies denotew where the R_ and S_ columns are in this particular file."""
        maps = self._create_maps(q_row, indices['response_indices'], indices['score_indices'])
        return Question.from_questions_rows(q_row, maps['response_options'], maps['score_map'])

    def _create_historic_questions(self, question):
        """For all historic survey periods, add HistoricQuestion object to the exisiting current period question.

        If config.HEAD_PDIFF column is one, denotes that the responses or scoring is diofferent form the current period,
        so recaulte info. Otherwise, scoring information is inherited from current period question.

        """
        for period, historic_df in self._historic_dfs.items():
            h_row = historic_df.loc[historic_df[config.HEAD_QVAR] == question.qid].squeeze()
            historic_qid = h_row[config.HEAD_PQVAR]

            if isinstance(historic_qid, str):
                # If PDIFF is 1, implies scoring/responses are different from current period sp
                if h_row[config.HEAD_PDIFF] == 1:
                    maps = self._create_maps(h_row, self._historic_indicies[period]['response_indices'],
                                             self._historic_indicies[period]['score_indices'])
                    question.add_historic_question(period, historic_qid, maps['response_options'], maps['score_map'])
                else:
                    question.add_historic_question(period, historic_qid)

    def _create_type_lists(self, q_type):
        """Generic function for creating lists based on question properties"""
        return [q for q in self._question_list if q.q_type == q_type]

    def get_by_qid(self, qid):
        """returns question object by QID"""
        return next((q for q in self._question_list if q.qid == qid), None)

    def __iter__(self):
        return iter(self._question_list)

    def __len__(self):
        return len(self._question_list)
